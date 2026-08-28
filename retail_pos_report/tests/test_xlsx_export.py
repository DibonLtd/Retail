import base64
import io

import openpyxl

from odoo.exceptions import UserError
from odoo.tests import tagged

from .common import PosReportCase


@tagged("post_install", "-at_install")
class TestXlsxExport(PosReportCase):
    """US-RPT-02: export the summary to Excel.

    The workbook is read back with openpyxl rather than merely checked for
    being non-empty, so the sheets and the grand-total rows are genuinely
    asserted.
    """

    def setUp(self):
        super().setUp()
        session = self._open_session()
        # 2 x unga (240) + 3 x milk (195) = 435, split cash and M-PESA.
        self._make_order(
            session,
            [(self.unga, 2), (self.milk, 3)],
            [(self.method_cash, 200.0), (self.method_mpesa, 235.0)],
            order_date="2025-01-24 09:00:00",
        )
        self._make_order(
            session,
            [(self.unga, 1)],
            [(self.method_cash, 120.0)],
            order_date="2025-01-21 09:00:00",
        )
        self.report = self._new_report("2025-01-20", "2025-01-26")
        self.report.action_generate()

    def _workbook(self):
        self.report.action_export_xlsx()
        raw = base64.b64decode(self.report.xlsx_file)
        return openpyxl.load_workbook(io.BytesIO(raw))

    def test_export_requires_a_generated_report(self):
        draft = self._new_report("2025-01-20", "2025-01-26")
        with self.assertRaises(UserError):
            draft.action_export_xlsx()

    def test_export_returns_a_download_action(self):
        action = self.report.action_export_xlsx()
        self.assertEqual(action["type"], "ir.actions.act_url")
        self.assertIn("download=true", action["url"])

    def test_filename_is_derived_from_the_reference(self):
        self.report.action_export_xlsx()
        self.assertTrue(self.report.xlsx_filename.endswith(".xlsx"))
        self.assertNotIn(
            "/", self.report.xlsx_filename, "Slashes would break the download."
        )

    def test_workbook_has_both_sheets(self):
        workbook = self._workbook()
        self.assertEqual(
            workbook.sheetnames, ["Daily Summary", "Sales by Item"]
        )

    def test_daily_sheet_headers(self):
        sheet = self._workbook()["Daily Summary"]
        headers = [cell.value for cell in sheet[4]]
        self.assertEqual(
            headers, ["Date", "Total Sales", "MPESA", "Cash", "Other", "Rank"]
        )

    def test_item_sheet_headers(self):
        sheet = self._workbook()["Sales by Item"]
        headers = [cell.value for cell in sheet[4]]
        self.assertEqual(
            headers,
            [
                "Dept",
                "Category",
                "Product",
                "Barcode",
                "Qty Sold",
                "Unit Price",
                "Total Sales",
                "Tax Ex",
                "VAT",
            ],
        )

    def test_daily_sheet_grand_total_row(self):
        sheet = self._workbook()["Daily Summary"]
        rows = list(sheet.iter_rows(values_only=True))
        total_row = next(row for row in rows if row and row[0] == "Grand Total")
        self.assertEqual(total_row[1], self.report.grand_total_sales)
        self.assertEqual(total_row[2], self.report.grand_total_mpesa)
        self.assertEqual(total_row[3], self.report.grand_total_cash)

    def test_daily_sheet_has_one_row_per_trading_day(self):
        sheet = self._workbook()["Daily Summary"]
        rows = list(sheet.iter_rows(min_row=5, values_only=True))
        data_rows = [row for row in rows if row[0] not in (None, "Grand Total")]
        self.assertEqual(len(data_rows), len(self.report.daily_line_ids))

    def test_item_sheet_grand_total_sums_the_lines(self):
        sheet = self._workbook()["Sales by Item"]
        rows = list(sheet.iter_rows(values_only=True))
        total_row = next(row for row in rows if row and row[0] == "Grand Total")
        self.assertEqual(
            total_row[6], sum(self.report.item_line_ids.mapped("total_sales"))
        )
        self.assertEqual(
            total_row[4], sum(self.report.item_line_ids.mapped("qty_sold"))
        )

    def test_item_sheet_carries_department_and_barcode(self):
        sheet = self._workbook()["Sales by Item"]
        rows = list(sheet.iter_rows(min_row=5, values_only=True))
        unga_row = next(row for row in rows if row[3] == "6001253001001")
        self.assertEqual(unga_row[0], "Dry Foods")
        self.assertEqual(unga_row[1], "Flour & Grains")

    def test_regenerating_refreshes_the_export(self):
        first = self._workbook()
        self.report.action_generate()
        second = self._workbook()
        self.assertEqual(first.sheetnames, second.sheetnames)
        self.assertTrue(self.report.xlsx_file)
